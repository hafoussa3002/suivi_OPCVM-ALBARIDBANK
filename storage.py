from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE_DATA_DIR = Path("data")
ASFIM_BASE_DIR = BASE_DATA_DIR / "asfim"
ASFIM_DIRS = {
    "quotidien": ASFIM_BASE_DIR / "quotidien",
    "hebdomadaire": ASFIM_BASE_DIR / "hebdomadaire",
}
BAM_BASE_DIR = BASE_DATA_DIR / "bam"
DB_DIR = BASE_DATA_DIR / "db"
HISTORY_PATH = DB_DIR / "history.json"


def init_storage() -> None:
    BASE_DATA_DIR.mkdir(parents=True, exist_ok=True)
    ASFIM_BASE_DIR.mkdir(parents=True, exist_ok=True)
    BAM_BASE_DIR.mkdir(parents=True, exist_ok=True)
    DB_DIR.mkdir(parents=True, exist_ok=True)
    for folder in ASFIM_DIRS.values():
        folder.mkdir(parents=True, exist_ok=True)
    if not HISTORY_PATH.exists():
        HISTORY_PATH.write_text(
            json.dumps({"version": 1, "items": []}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def load_history() -> dict[str, Any]:
    init_storage()
    try:
        payload = json.loads(HISTORY_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        payload = {"version": 1, "items": []}
    if "items" not in payload or not isinstance(payload["items"], list):
        payload["items"] = []
    if "version" not in payload:
        payload["version"] = 1
    return payload


def save_history(history: dict[str, Any]) -> None:
    BASE_DATA_DIR.mkdir(parents=True, exist_ok=True)
    ASFIM_BASE_DIR.mkdir(parents=True, exist_ok=True)
    BAM_BASE_DIR.mkdir(parents=True, exist_ok=True)
    DB_DIR.mkdir(parents=True, exist_ok=True)
    for folder in ASFIM_DIRS.values():
        folder.mkdir(parents=True, exist_ok=True)
    HISTORY_PATH.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_frequency(frequency: str) -> str:
    value = frequency.strip().lower()
    if value not in ASFIM_DIRS:
        raise ValueError("frequency doit être 'quotidien' ou 'hebdomadaire'")
    return value


def _sanitize_date_key(date_key: str) -> str:
    clean = re.sub(r"[^0-9A-Za-z_-]", "-", date_key.strip())
    clean = re.sub(r"-+", "-", clean).strip("-")
    if not clean:
        raise ValueError("date_key invalide")
    return clean


def _sanitize_filename(filename: str) -> str:
    safe = re.sub(r"[^0-9A-Za-z._-]", "_", filename)
    safe = re.sub(r"_+", "_", safe)
    return safe.strip("_") or "asfim.xlsx"


def _sort_date_keys(keys: list[str]) -> list[str]:
    def parse_key(k: str) -> tuple[int, datetime | None, str]:
        for fmt in ("%Y-%m-%d", "%Y_%m_%d", "%d-%m-%Y", "%d_%m_%Y"):
            try:
                return (0, datetime.strptime(k, fmt), k)
            except ValueError:
                continue
        return (1, None, k)

    parsed = [parse_key(k) for k in keys]
    dated = sorted([p for p in parsed if p[0] == 0], key=lambda x: x[1], reverse=True)
    raw = sorted([p for p in parsed if p[0] == 1], key=lambda x: x[2], reverse=True)
    return [p[2] for p in dated + raw]


def _extract_date_from_filename(filename: str) -> str | None:
    patterns = [
        r"(\d{4}-\d{2}-\d{2})",
        r"(\d{2}-\d{2}-\d{4})",
        r"(\d{2}_\d{2}_\d{4})",
        r"(\d{4}_\d{2}_\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            candidate = match.group(1).replace("_", "-")
            try:
                if re.match(r"\d{2}-\d{2}-\d{4}", candidate):
                    return datetime.strptime(candidate, "%d-%m-%Y").strftime("%Y-%m-%d")
                return datetime.strptime(candidate, "%Y-%m-%d").strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def _extract_date_from_content(file_bytes: bytes) -> str | None:
    wb = None
    try:
        from io import BytesIO

        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        max_rows = min(6, ws.max_row)
        for row_idx in range(1, max_rows + 1):
            values = ws[row_idx]
            text = " ".join(str(c.value) for c in values if c.value is not None)
            match = re.search(r"(\d{2}[/-]\d{2}[/-]\d{4})", text)
            if match:
                raw = match.group(1).replace("-", "/")
                parsed = datetime.strptime(raw, "%d/%m/%Y")
                return parsed.strftime("%Y-%m-%d")
    except Exception:
        return None
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass

    return None


def _extract_date_from_text(text: str) -> str | None:
    if not text:
        return None
    m1 = re.search(r"(\d{2}[/-]\d{2}[/-]\d{4})", text)
    if m1:
        raw = m1.group(1).replace("-", "/")
        try:
            return datetime.strptime(raw, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass
    m2 = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    if m2:
        try:
            return datetime.strptime(m2.group(1), "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def _extract_bam_date_from_content(file_bytes: bytes) -> str | None:
    wb = None
    try:
        from io import BytesIO

        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        header_row = None
        date_col = None
        for r in range(1, min(15, ws.max_row) + 1):
            row_vals = ["" if c.value is None else str(c.value) for c in ws[r]]
            for i, v in enumerate(row_vals, start=1):
                norm = v.strip().lower()
                if norm in {"date de la valeur", "date de valeur"}:
                    header_row = r
                    date_col = i
                    break
            if header_row:
                break

        if header_row and date_col:
            counts: dict[str, int] = {}
            for r in range(header_row + 1, min(ws.max_row, header_row + 300) + 1):
                val = ws.cell(row=r, column=date_col).value
                parsed = _extract_date_from_text("" if val is None else str(val))
                if parsed:
                    counts[parsed] = counts.get(parsed, 0) + 1
            if counts:
                return sorted(counts.items(), key=lambda x: (-x[1], x[0]))[0][0]

        for r in range(1, min(8, ws.max_row) + 1):
            text = " ".join("" if c.value is None else str(c.value) for c in ws[r])
            parsed = _extract_date_from_text(text)
            if parsed:
                return parsed
    except Exception:
        return None
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass
    return None


def _resolve_date_key(uploaded_file, batch_date_key: str | None = None, kind: str = "asfim") -> tuple[str | None, str]:
    date_from_content = _extract_date_from_content(uploaded_file.getvalue())
    if kind == "bam":
        date_from_content = _extract_bam_date_from_content(uploaded_file.getvalue()) or date_from_content
    if date_from_content:
        return date_from_content, "contenu"

    date_from_filename = _extract_date_from_filename(uploaded_file.name)
    if date_from_filename:
        return date_from_filename, "nom_fichier"

    if batch_date_key:
        return _sanitize_date_key(batch_date_key), "date_lot"

    return None, "manquante"


def _build_storage_path(frequency: str, date_key: str, original_filename: str) -> Path:
    folder = ASFIM_DIRS[frequency]
    return _build_storage_path_in_folder(folder, date_key, original_filename)


def _build_storage_path_in_folder(folder: Path, date_key: str, original_filename: str) -> Path:
    safe_date = _sanitize_date_key(date_key)
    safe_name = _sanitize_filename(original_filename)

    candidate = folder / f"{safe_date}__{safe_name}"
    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix or ".xlsx"
    version = 2
    while True:
        versioned = folder / f"{stem}_v{version}{suffix}"
        if not versioned.exists():
            return versioned
        version += 1


def add_asfim_files(files, frequency: str, batch_date_key: str | None = None) -> dict[str, Any]:
    init_storage()
    normalized_frequency = _normalize_frequency(frequency)

    history = load_history()
    items: list[dict[str, Any]] = history.get("items", [])

    results = {"saved": [], "errors": []}
    for f in files:
        resolved_date_key, source = _resolve_date_key(f, batch_date_key=batch_date_key, kind="asfim")
        if not resolved_date_key:
            results["errors"].append(
                {
                    "filename": f.name,
                    "error": "Impossible de déterminer date_key (contenu/nom/date lot).",
                }
            )
            continue

        target_path = _build_storage_path(normalized_frequency, resolved_date_key, f.name)
        target_path.write_bytes(f.getvalue())

        record = {
            "kind": "asfim",
            "frequency": normalized_frequency,
            "date_key": _sanitize_date_key(resolved_date_key),
            "filename": target_path.name,
            "original_filename": f.name,
            "storage_path": str(target_path.as_posix()),
            "uploaded_at": datetime.now().isoformat(timespec="seconds"),
            "date_source": source,
        }
        items.append(record)
        results["saved"].append(record)

    history["items"] = items
    save_history(history)
    return results


def list_asfim_dates(frequency: str) -> list[str]:
    normalized_frequency = _normalize_frequency(frequency)
    items = load_history().get("items", [])
    dates = sorted(
        {
            str(item.get("date_key", ""))
            for item in items
            if item.get("kind") == "asfim" and item.get("frequency") == normalized_frequency and item.get("date_key")
        }
    )
    return _sort_date_keys(dates)


def list_asfim_files(frequency: str, date_key: str) -> list[dict[str, Any]]:
    normalized_frequency = _normalize_frequency(frequency)
    normalized_date_key = _sanitize_date_key(date_key)
    items = load_history().get("items", [])
    out = [
        item
        for item in items
        if item.get("kind") == "asfim"
        and item.get("frequency") == normalized_frequency
        and item.get("date_key") == normalized_date_key
    ]
    out.sort(key=lambda x: str(x.get("uploaded_at", "")), reverse=True)
    return out


def summarize_asfim_history() -> list[dict[str, Any]]:
    items = [i for i in load_history().get("items", []) if i.get("kind") == "asfim"]
    summary: dict[tuple[str, str], int] = {}
    for it in items:
        key = (str(it.get("frequency", "")), str(it.get("date_key", "")))
        summary[key] = summary.get(key, 0) + 1

    rows = [
        {"Type": freq, "Date": date_key, "Nombre de fichiers": count}
        for (freq, date_key), count in summary.items()
        if freq and date_key
    ]

    rows.sort(key=lambda r: (r["Type"], r["Date"]), reverse=True)
    return rows


def get_asfim_records(frequency: str | None = None, date_key: str | None = None) -> list[dict[str, Any]]:
    items = [i for i in load_history().get("items", []) if i.get("kind") == "asfim"]
    if frequency:
        normalized_frequency = _normalize_frequency(frequency)
        items = [i for i in items if i.get("frequency") == normalized_frequency]
    if date_key:
        normalized_date = _sanitize_date_key(date_key)
        items = [i for i in items if i.get("date_key") == normalized_date]
    items.sort(key=lambda x: str(x.get("uploaded_at", "")), reverse=True)
    return items


def add_bam_files(files, batch_date_key: str | None = None) -> dict[str, Any]:
    init_storage()
    history = load_history()
    items: list[dict[str, Any]] = history.get("items", [])
    results = {"saved": [], "errors": []}

    for f in files:
        resolved_date_key, source = _resolve_date_key(f, batch_date_key=batch_date_key, kind="bam")
        if not resolved_date_key:
            results["errors"].append(
                {
                    "filename": f.name,
                    "error": "Impossible de déterminer date_key BAM (contenu/nom/date lot).",
                }
            )
            continue

        target_path = _build_storage_path_in_folder(BAM_BASE_DIR, resolved_date_key, f.name)
        target_path.write_bytes(f.getvalue())

        record = {
            "kind": "bam",
            "date_key": _sanitize_date_key(resolved_date_key),
            "filename": target_path.name,
            "original_filename": f.name,
            "storage_path": str(target_path.as_posix()),
            "uploaded_at": datetime.now().isoformat(timespec="seconds"),
            "date_source": source,
        }
        items.append(record)
        results["saved"].append(record)

    history["items"] = items
    save_history(history)
    return results


def list_bam_dates() -> list[str]:
    items = load_history().get("items", [])
    dates = sorted({str(i.get("date_key", "")) for i in items if i.get("kind") == "bam" and i.get("date_key")})
    return _sort_date_keys(dates)


def list_bam_files(date_key: str) -> list[dict[str, Any]]:
    normalized_date_key = _sanitize_date_key(date_key)
    items = load_history().get("items", [])
    out = [i for i in items if i.get("kind") == "bam" and i.get("date_key") == normalized_date_key]
    out.sort(key=lambda x: str(x.get("uploaded_at", "")), reverse=True)
    return out


def summarize_bam_history() -> list[dict[str, Any]]:
    items = [i for i in load_history().get("items", []) if i.get("kind") == "bam"]
    summary: dict[str, int] = {}
    for it in items:
        date_key = str(it.get("date_key", ""))
        if date_key:
            summary[date_key] = summary.get(date_key, 0) + 1
    rows = [{"Date": k, "Nombre de fichiers": v} for k, v in summary.items()]
    rows.sort(key=lambda r: r["Date"], reverse=True)
    return rows


def get_bam_records(date_key: str | None = None) -> list[dict[str, Any]]:
    items = [i for i in load_history().get("items", []) if i.get("kind") == "bam"]
    if date_key:
        normalized_date_key = _sanitize_date_key(date_key)
        items = [i for i in items if i.get("date_key") == normalized_date_key]
    items.sort(key=lambda x: str(x.get("uploaded_at", "")), reverse=True)
    return items
